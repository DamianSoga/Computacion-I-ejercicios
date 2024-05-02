import openpyxl
import matplotlib.pyplot as plt

wb = openpyxl.load_workbook('/home/alumno/Downloads/videogamesales.xlsx')
ws = wb.active
print(ws['A13'].value)

#filas = [ws.cell(row=1,column=i).value for i in range(1,10)]
#print(filas)
years = [ws.cell(row=i,column=4).value for i in range(2,16329)]
miny = min(years)
maxy = max(years)

def eval(iter,yearA,yearB):
    if iter>=yearA and iter<=yearB:
        return iter

catA = [eval(y,1980,1990) for y in years]
catAf = list(filter(lambda x: x!=None,catA))

plt.hist(years)

plt.show()